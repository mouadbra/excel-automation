from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import os

class ExcelManager:
    def __init__(self, filename):
        self.filename = filename
        self.workbook = None
        self.worksheet = None
    
    def create_new_workbook(self):
        """Create a new Excel workbook with headers"""
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = "Weather Data"
        
        # Headers
        headers = [
            'Timestamp', 'City', 'Country', 'Temperature (°C)', 
            'Feels Like (°C)', 'Humidity (%)', 'Pressure (hPa)', 
            'Weather', 'Description', 'Wind Speed (m/s)'
        ]
        
        self.worksheet.append(headers)
        self._format_headers()
        self.save_workbook()
    
    def load_workbook(self):
        """Load existing workbook or create new one"""
        try:
            self.workbook = load_workbook(self.filename)
            self.worksheet = self.workbook.active
        except FileNotFoundError:
            print(f"File {self.filename} not found. Creating new workbook.")
            self.create_new_workbook()
    
    def _format_headers(self):
        """Format header row"""
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col in range(1, 11):  # 10 columns
            cell = self.worksheet.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Auto-adjust column widths
        for col in range(1, 11):
            column_letter = get_column_letter(col)
            self.worksheet.column_dimensions[column_letter].width = 15
    
    def add_weather_data(self, weather_data):
        """Add weather data to the worksheet"""
        for data in weather_data:
            row_data = [
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                data['city'],
                data['country'],
                data['temperature'],
                data['feels_like'],
                data['humidity'],
                data['pressure'],
                data['weather'],
                data['description'],
                data['wind_speed']
            ]
            self.worksheet.append(row_data)
    
    def clear_old_data(self, keep_days=7):
        """Clear data older than specified days"""
        if self.worksheet.max_row <= 1:
            return
        
        cutoff_date = datetime.now().timestamp() - (keep_days * 24 * 3600)
        rows_to_delete = []
        
        for row in range(2, self.worksheet.max_row + 1):
            timestamp_cell = self.worksheet.cell(row=row, column=1)
            if timestamp_cell.value:
                try:
                    row_timestamp = datetime.strptime(
                        str(timestamp_cell.value), "%Y-%m-%d %H:%M:%S"
                    ).timestamp()
                    if row_timestamp < cutoff_date:
                        rows_to_delete.append(row)
                except (ValueError, TypeError):
                    continue
        
        # Delete rows in reverse order to maintain indices
        for row in reversed(rows_to_delete):
            self.worksheet.delete_rows(row)
    
    def create_summary_sheet(self):
        """Create a summary sheet with statistics"""
        if "Summary" in self.workbook.sheetnames:
            del self.workbook["Summary"]
        
        summary_sheet = self.workbook.create_sheet("Summary")
        
        # Add summary headers
        summary_sheet.append(["City", "Avg Temperature", "Max Temperature", "Min Temperature", "Last Update"])
        
        # Get unique cities from main sheet
        cities = set()
        for row in range(2, self.worksheet.max_row + 1):
            city = self.worksheet.cell(row=row, column=2).value
            if city:
                cities.add(city)
        
        # Calculate statistics for each city
        for city in cities:
            temps = []
            last_update = ""
            
            for row in range(2, self.worksheet.max_row + 1):
                if self.worksheet.cell(row=row, column=2).value == city:
                    temp = self.worksheet.cell(row=row, column=4).value
                    timestamp = self.worksheet.cell(row=row, column=1).value
                    
                    if temp is not None:
                        temps.append(float(temp))
                    if timestamp:
                        last_update = timestamp
            
            if temps:
                avg_temp = sum(temps) / len(temps)
                max_temp = max(temps)
                min_temp = min(temps)
                
                summary_sheet.append([
                    city,
                    round(avg_temp, 1),
                    round(max_temp, 1),
                    round(min_temp, 1),
                    last_update
                ])
    
    def save_workbook(self):
        """Save the workbook"""
        if self.workbook:
            self.workbook.save(self.filename)
            print(f"Workbook saved: {self.filename}")
