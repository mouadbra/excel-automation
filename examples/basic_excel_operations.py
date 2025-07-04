
# === examples/basic_excel_operations.py ===
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

def create_basic_excel():
    """Example: Creating a basic Excel file"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    
    # Add data
    ws.append(['Name', 'Age', 'City', 'Score'])
    ws.append(['Alice', 25, 'Paris', 85])
    ws.append(['Bob', 30, 'London', 92])
    ws.append(['Charlie', 35, 'New York', 78])
    
    wb.save('basic_example.xlsx')
    print("Basic Excel file created: basic_example.xlsx")

def modify_excel_sheet():
    """Example: Modifying an existing Excel sheet"""
    wb = load_workbook('basic_example.xlsx')
    ws = wb.active
    
    # Modify a specific cell
    ws['A1'].value = "Full Name"
    ws['A1'].font = Font(bold=True)
    
    # Add a new column
    ws['E1'] = "Grade"
    ws['E1'].font = Font(bold=True)
    
    # Add grades based on scores
    for row in range(2, ws.max_row + 1):
        score = ws[f'D{row}'].value
        if score >= 90:
            ws[f'E{row}'] = 'A'
        elif score >= 80:
            ws[f'E{row}'] = 'B'
        elif score >= 70:
            ws[f'E{row}'] = 'C'
        else:
            ws[f'E{row}'] = 'D'
    
    wb.save('basic_example.xlsx')
    print("Excel file modified successfully")

def merge_cells_example():
    """Example: Merging and formatting cells"""
    wb = load_workbook('basic_example.xlsx')
    ws = wb.active
    
    # Insert a new row at the top
    ws.insert_rows(1)
    
    # Merge cells for title
    ws.merge_cells("A1:E1")
    ws['A1'] = "Student Data Report"
    ws['A1'].font = Font(bold=True, size=16)
    
    wb.save('basic_example.xlsx')
    print("Cells merged and formatted")

if __name__ == "__main__":
    create_basic_excel()
    modify_excel_sheet()
    merge_cells_example()